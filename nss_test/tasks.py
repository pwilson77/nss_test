from datamodel.models import Datamodel
import openpyxl
from celery import shared_task
from celery.utils.log import get_task_logger
import os

logger = get_task_logger(__name__)


@shared_task
def upload_data(excel_file):

    # validations can be placed here to check file extension
    #exc = open(excel_file, "rb")
    # logger.info(excel_file)
    wb = openpyxl.load_workbook(excel_file)

    # getting a particular sheet by name out of many sheets

    worksheet = wb.active

    # iterating over the rows and
    # getting value from each cell in row

    for i in range(2, worksheet.max_row + 1):
        row_data = Datamodel()
        #logger.info(worksheet.cell(row=i, column=3).value)
        row_data.firstname = worksheet.cell(row=i, column=1).value
        row_data.lastname = worksheet.cell(row=i, column=2).value
        row_data.age = worksheet.cell(row=i, column=3).value
        row_data.gender = worksheet.cell(row=i, column=4).value
        row_data.address = worksheet.cell(row=i, column=5).value

        row_data.save()

    os.remove(excel_file)
